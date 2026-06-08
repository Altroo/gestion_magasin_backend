from datetime import date, time
from io import BytesIO
from decimal import Decimal
from unittest.mock import patch

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from attendance.models import AttendanceRecord, Employee
from store.models import Role, Store, StoreMembership

pytestmark = pytest.mark.django_db

User = get_user_model()


def authenticated_client(user):
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def create_store_setup():
    user = User.objects.create_user(email="attendance@example.com", password="securepass123")
    role, _ = Role.objects.get_or_create(
        code=Role.Codes.RESPONSABLE,
        defaults={"name": "Responsable", "rank": 1},
    )
    store = Store.objects.create(code="attendance-store", name="ATTENDANCE STORE", is_active=True)
    StoreMembership.objects.create(user=user, store=store, role=role)
    employee = Employee.objects.create(store=store, full_name="Employé test")
    return user, store, employee


def test_attendance_list_accepts_comma_separated_status_filters():
    user, store, employee = create_store_setup()
    present = AttendanceRecord.objects.create(
        store=store,
        employee=employee,
        date=date(2026, 6, 1),
        hours_worked=Decimal("8.00"),
        status=AttendanceRecord.Statuses.PRESENT,
    )
    off = AttendanceRecord.objects.create(
        store=store,
        employee=employee,
        date=date(2026, 6, 2),
        hours_worked=Decimal("0.00"),
        status=AttendanceRecord.Statuses.OFF,
    )
    AttendanceRecord.objects.create(
        store=store,
        employee=employee,
        date=date(2026, 6, 3),
        hours_worked=Decimal("0.00"),
        status=AttendanceRecord.Statuses.ABSENT,
    )
    client = authenticated_client(user)

    response = client.get(
        "/api/pointage/",
        {
            "store": store.pk,
            "status": "present,off",
        },
    )

    assert response.status_code == status.HTTP_200_OK
    assert {item["id"] for item in response.data["results"]} == {present.pk, off.pk}


def test_attendance_list_filters_by_store_ids_and_employee_ids():
    user, store, employee = create_store_setup()
    role = StoreMembership.objects.get(user=user, store=store).role
    second_store = Store.objects.create(code="attendance-second", name="SECOND STORE", is_active=True)
    StoreMembership.objects.create(user=user, store=second_store, role=role)
    second_employee = Employee.objects.create(store=second_store, full_name="Second employé")
    third_employee = Employee.objects.create(store=store, full_name="Autre employé")
    matching = AttendanceRecord.objects.create(
        store=second_store,
        employee=second_employee,
        date=date(2026, 6, 4),
        hours_worked=Decimal("7.50"),
        status=AttendanceRecord.Statuses.PRESENT,
    )
    AttendanceRecord.objects.create(
        store=store,
        employee=employee,
        date=date(2026, 6, 5),
        hours_worked=Decimal("8.00"),
        status=AttendanceRecord.Statuses.PRESENT,
    )
    AttendanceRecord.objects.create(
        store=store,
        employee=third_employee,
        date=date(2026, 6, 6),
        hours_worked=Decimal("8.00"),
        status=AttendanceRecord.Statuses.PRESENT,
    )
    client = authenticated_client(user)

    response = client.get(
        "/api/pointage/",
        {
            "store_ids": f"{store.pk},{second_store.pk}",
            "employee_ids": str(second_employee.pk),
        },
    )

    assert response.status_code == status.HTTP_200_OK
    assert response.data["count"] == 1
    assert response.data["results"][0]["id"] == matching.pk


def test_attendance_create_calculates_hours_and_delay():
    user, store, employee = create_store_setup()
    client = authenticated_client(user)

    response = client.post(
        "/api/pointage/",
        {
            "store": store.pk,
            "employee": employee.pk,
            "date": "2026-06-08",
            "clock_in": "09:20",
            "break_start": "13:30",
            "break_end": "14:00",
            "clock_out": "17:00",
            "shift": AttendanceRecord.Shifts.MORNING,
            "hours_worked": "99.00",
            "delay_minutes": 999,
            "status": AttendanceRecord.Statuses.PRESENT,
            "responsible": "Responsable",
            "observations": "",
        },
        format="json",
    )

    assert response.status_code == status.HTTP_201_CREATED
    record = AttendanceRecord.objects.get(pk=response.data["id"])
    assert record.hours_worked == Decimal("7.17")
    assert record.delay_minutes == 20


def test_attendance_bulk_delete_removes_selected_records():
    user, store, employee = create_store_setup()
    first = AttendanceRecord.objects.create(
        store=store,
        employee=employee,
        date=date(2026, 6, 10),
        hours_worked=Decimal("8.00"),
        status=AttendanceRecord.Statuses.PRESENT,
    )
    second = AttendanceRecord.objects.create(
        store=store,
        employee=employee,
        date=date(2026, 6, 11),
        hours_worked=Decimal("8.00"),
        status=AttendanceRecord.Statuses.PRESENT,
    )
    client = authenticated_client(user)

    response = client.delete(
        "/api/pointage/bulk-delete/",
        {"ids": [first.pk, second.pk]},
        format="json",
    )

    assert response.status_code == status.HTTP_200_OK
    assert response.data["deleted"] == 2
    assert not AttendanceRecord.objects.filter(pk__in=[first.pk, second.pk]).exists()


def test_attendance_export_workbook_matches_import_layout():
    user, store, employee = create_store_setup()
    AttendanceRecord.objects.create(
        store=store,
        employee=employee,
        date=date(2026, 6, 10),
        clock_in=time(9, 0),
        break_start=time(13, 30),
        break_end=time(14, 0),
        clock_out=time(17, 0),
        hours_worked=Decimal("7.50"),
        status=AttendanceRecord.Statuses.PRESENT,
        shift=AttendanceRecord.Shifts.MORNING,
    )
    client = authenticated_client(user)

    response = client.get("/api/pointage/export-workbook/", {"store": store.pk})

    assert response.status_code == status.HTTP_200_OK
    assert response["Content-Disposition"] == 'attachment; filename="pointage.xlsx"'
    workbook = load_workbook(BytesIO(b"".join(response.streaming_content) if getattr(response, "streaming", False) else response.content))
    sheet = workbook["Pointage"]
    assert sheet["A2"].value == "FICHE DE POINTAGE HEBDOMADAIRE -MBR SOUTH"
    assert [sheet.cell(5, column).value for column in range(1, 12)] == [
        "Date",
        "Jour",
        "Nom salarié",
        "Heure entrée",
        "Début pause",
        "Fin pause",
        "Heure sortie",
        "Statut",
        "Shift",
        "Retard",
        "Observations",
    ]
    assert sheet["C6"].value == "Employé test"
    assert sheet["D6"].value == "9H"


def test_attendance_import_guide_email_requires_management_access():
    user, store, _employee = create_store_setup()
    StoreMembership.objects.filter(user=user, store=store).delete()
    client = authenticated_client(user)

    response = client.post("/api/pointage/send-import-guide-email/", {"store": store.pk}, format="json")

    assert response.status_code == status.HTTP_403_FORBIDDEN


def test_attendance_import_guide_email_schedules_email():
    user, store, _employee = create_store_setup()
    client = authenticated_client(user)

    with patch("account.tasks.send_attendance_import_guide_email.apply_async") as mocked_task:
        response = client.post("/api/pointage/send-import-guide-email/", {"store": store.pk}, format="json")

    assert response.status_code == status.HTTP_200_OK
    mocked_task.assert_called_once_with((user.pk, user.email))


def test_attendance_model_calculates_evening_shift_delay():
    user, store, employee = create_store_setup()

    record = AttendanceRecord.objects.create(
        store=store,
        employee=employee,
        date=date(2026, 6, 9),
        clock_in=time(15, 12),
        clock_out=time(23, 0),
        shift=AttendanceRecord.Shifts.EVENING,
        status=AttendanceRecord.Statuses.PRESENT,
        created_by=user,
    )

    assert record.hours_worked == Decimal("7.80")
    assert record.delay_minutes == 12
