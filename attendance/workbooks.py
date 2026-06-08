from datetime import date, time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


POINTAGE_HEADERS = [
    "Date",
    "Jour",
    "Nom salarié",
    "Heure  entrée",
    "Début pause",
    "Fin pause",
    "Heure sortie",
    "Statut",
    "Shift",
    "Retard",
    "Observations",
]

POINTAGE_COLUMN_WIDTHS = {
    "A": 13.21875,
    "B": 15.44140625,
    "C": 16.88671875,
    "D": 11.5546875,
    "E": 10.44140625,
    "F": 10.21875,
    "G": 9.6640625,
    "H": 13,
    "I": 13,
    "J": 13,
    "K": 25,
}

FRENCH_DAYS = {
    0: "Lundi",
    1: "Mardi",
    2: "Mercredi",
    3: "Jeudi",
    4: "Vendredi",
    5: "Samedi",
    6: "Dimanche",
}


def _format_time(value: time | None) -> str:
    if not value:
        return ""
    return f"{value.hour}H{value.minute:02d}" if value.minute else f"{value.hour}H"


def _status_label(status: str | None) -> str:
    if status == "off":
        return "OFF"
    if status == "absent":
        return "ABSENT"
    return "ACTIVE"


def _shift_label(shift: str | None) -> str:
    if shift == "evening":
        return "EVENING"
    if shift == "off":
        return "OFF"
    return "MORNING"


def _day_label(value: date | None) -> str:
    return FRENCH_DAYS[value.weekday()] if value else ""


def _style_sheet(sheet):
    title_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
    header_fill = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=False)
    header_font = Font(name="Calibri", size=12, bold=False)
    data_font = Font(name="Calibri", size=11, bold=False)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    medium_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    sheet.sheet_view.showGridLines = None
    sheet.row_dimensions[1].height = 15
    sheet.row_dimensions[3].height = 15
    sheet.row_dimensions[4].height = 25.2
    sheet.row_dimensions[5].height = 33
    for column_letter, width in POINTAGE_COLUMN_WIDTHS.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.merge_cells("A2:K3")
    for row in sheet.iter_rows(min_row=2, max_row=3, min_col=1, max_col=11):
        for cell in row:
            cell.fill = title_fill
            cell.border = medium_border
    sheet["A2"].font = title_font
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[5]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=6):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="left")
                cell.number_format = "mm-dd-yy"
            elif cell.column < 11:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.auto_filter.ref = f"A5:K{max(sheet.max_row, 5)}"


def build_attendance_workbook(records, *, responsible="", week_start=None, week_end=None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pointage"

    sheet.append([None] * len(POINTAGE_HEADERS))
    sheet.append(["FICHE DE POINTAGE HEBDOMADAIRE-MBR SOUTH", *([None] * 10)])
    sheet.append([None] * len(POINTAGE_HEADERS))
    sheet.append([None] * len(POINTAGE_HEADERS))
    sheet.append(POINTAGE_HEADERS)

    for record in records:
        current_date = record.date
        clock_in = "OFF" if record.status == "off" else _format_time(record.clock_in)
        sheet.append([
            current_date,
            _day_label(current_date),
            record.employee.full_name,
            clock_in,
            _format_time(record.break_start),
            _format_time(record.break_end),
            _format_time(record.clock_out),
            _status_label(record.status),
            _shift_label(record.shift),
            int(record.delay_minutes or 0),
            record.observations or None,
        ])

    _style_sheet(sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_attendance_import_template() -> bytes:
    class TemplateEmployee:
        full_name = "INES"

    class TemplateRecord:
        date = date.today()
        employee = TemplateEmployee()
        clock_in = time(9, 0)
        break_start = time(13, 30)
        break_end = time(14, 0)
        clock_out = time(17, 0)
        status = "present"
        shift = "morning"
        delay_minutes = 0
        observations = "NO"

    return build_attendance_workbook(
        [TemplateRecord()],
        responsible="Responsable",
        week_start=date.today(),
        week_end=date.today(),
    )
