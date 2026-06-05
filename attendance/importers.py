from datetime import date, datetime, time
from typing import BinaryIO

from django.db import transaction

from attendance.models import AttendanceImportBatch, AttendanceRecord, Employee


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _date_or_none(value) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _time_or_none(value) -> time | None:
    text = _clean(value).lower()
    if not text or text == "off":
        return None
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    text = text.replace(" ", "").replace("h", ":")
    if text.endswith(":"):
        text = f"{text}00"
    for fmt in ("%H:%M", "%H"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def _shift_or_default(value, status) -> str:
    text = _clean(value).lower()
    if status == AttendanceRecord.Statuses.OFF or text == "off":
        return AttendanceRecord.Shifts.OFF
    if "evening" in text or "soir" in text:
        return AttendanceRecord.Shifts.EVENING
    return AttendanceRecord.Shifts.MORNING


def _find_header_row(sheet) -> int:
    for row_number in range(1, min(sheet.max_row, 15) + 1):
        values = [_clean(cell.value).lower() for cell in sheet[row_number]]
        if "date" in values and any("nom" in value for value in values):
            return row_number
    return 5


@transaction.atomic
def import_attendance_from_workbook(
    file_obj: BinaryIO | str,
    *,
    store,
    imported_by=None,
    file_name: str = "",
) -> AttendanceImportBatch:
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook["Pointage"] if "Pointage" in workbook.sheetnames else workbook.active
    responsible = _clean(sheet["B3"].value) or _clean(sheet["C3"].value)
    week_start = _date_or_none(sheet["E3"].value)
    week_end = _date_or_none(sheet["G3"].value)
    header_row_number = _find_header_row(sheet)
    headers = [_clean(cell.value).lower() for cell in sheet[header_row_number]]

    def idx(*names):
        for name in names:
            for index, header in enumerate(headers):
                if name in header:
                    return index
        return None

    date_idx = idx("date")
    name_idx = idx("nom", "employ")
    in_idx = idx("entrée", "entree")
    break_start_idx = idx("début", "debut", "pause 1")
    break_end_idx = idx("fin pause", "pause 2")
    out_idx = idx("sortie")
    shift_idx = idx("shift")
    obs_idx = idx("observ")

    batch = AttendanceImportBatch.objects.create(
        store=store,
        file_name=file_name or getattr(file_obj, "name", "") or "pointage.xlsx",
        responsible=responsible,
        week_start=week_start,
        week_end=week_end,
        imported_by=imported_by if getattr(imported_by, "is_authenticated", False) else None,
    )

    imported = 0
    skipped = 0
    for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        current_date = _date_or_none(row[date_idx]) if date_idx is not None else None
        employee_name = _clean(row[name_idx]).upper() if name_idx is not None else ""
        if not current_date or not employee_name:
            skipped += 1
            continue
        raw_in = row[in_idx] if in_idx is not None else None
        clock_in = _time_or_none(raw_in)
        clock_out = _time_or_none(row[out_idx]) if out_idx is not None else None
        break_start = _time_or_none(row[break_start_idx]) if break_start_idx is not None else None
        break_end = _time_or_none(row[break_end_idx]) if break_end_idx is not None else None
        status = AttendanceRecord.Statuses.OFF if _clean(raw_in).lower() == "off" else AttendanceRecord.Statuses.PRESENT
        if status == AttendanceRecord.Statuses.PRESENT and not clock_in and not clock_out:
            status = AttendanceRecord.Statuses.ABSENT
        shift = _shift_or_default(row[shift_idx] if shift_idx is not None else None, status)

        employee, _ = Employee.objects.get_or_create(
            store=store,
            full_name=employee_name,
            defaults={"position": ""},
        )
        AttendanceRecord.objects.update_or_create(
            store=store,
            employee=employee,
            date=current_date,
            defaults={
                "clock_in": clock_in,
                "break_start": break_start,
                "break_end": break_end,
                "clock_out": clock_out,
                "shift": shift,
                "status": status,
                "responsible": responsible,
                "observations": _clean(row[obs_idx]) if obs_idx is not None else "",
                "created_by": imported_by if getattr(imported_by, "is_authenticated", False) else None,
            },
        )
        imported += 1

    batch.imported_count = imported
    batch.skipped_count = skipped
    batch.save(update_fields=["imported_count", "skipped_count"])
    return batch
