from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

from django.db import transaction

from catalog.models import Category, Product, ProductImportBatch
from stock.models import StockBalance


HEADER_MAP = {
    "Réf": "reference",
    "Ref": "reference",
    "Désignation": "name",
    "Designation": "name",
    "Famille": "category_code",
    "Unité Vente": "unit",
    "Unite Vente": "unit",
    "P Achat": "purchase_price",
    "P Gros": "wholesale_price",
    "P Details": "detail_price",
    "P Détails": "detail_price",
    "P Comptoir": "counter_price",
    "Stock Alert": "default_stock_alert",
    "Stock": "stock",
    "Date expiration": "expiration_date",
    "Duree": "shelf_life_days",
    "Durée": "shelf_life_days",
    "obligation": "compliance_required",
}


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _int_or_none(value) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _date_or_none(value) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _bool(value) -> bool:
    return _clean(value).lower() in {"oui", "yes", "true", "1", "obligatoire"}


def _category(code: str) -> Category | None:
    if not code:
        return None
    normalized = code.strip()
    return Category.objects.get_or_create(
        code=normalized,
        defaults={"name": f"Famille {normalized}"},
    )[0]


@transaction.atomic
def import_products_from_workbook(
    file_obj: BinaryIO | str,
    *,
    store,
    imported_by=None,
    file_name: str = "",
) -> ProductImportBatch:
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook["F"] if "F" in workbook.sheetnames else workbook.active
    header_row = [cell.value for cell in sheet[1]]
    indexes = {
        HEADER_MAP[_clean(header)]: idx
        for idx, header in enumerate(header_row)
        if _clean(header) in HEADER_MAP
    }

    batch = ProductImportBatch.objects.create(
        store=store,
        file_name=file_name or getattr(file_obj, "name", "") or "articles.xlsx",
        imported_by=imported_by if getattr(imported_by, "is_authenticated", False) else None,
    )

    imported = 0
    skipped = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        reference = _clean(row[indexes["reference"]]) if "reference" in indexes else ""
        name = _clean(row[indexes["name"]]) if "name" in indexes else ""
        if not reference and not name:
            skipped += 1
            continue

        category = _category(_clean(row[indexes["category_code"]])) if "category_code" in indexes else None
        unit = _clean(row[indexes["unit"]]) if "unit" in indexes else "unité"
        opening_stock = _decimal(row[indexes["stock"]]) if "stock" in indexes else Decimal("0")
        lookup = {"reference": reference} if reference else {"name": name}
        product, _ = Product.objects.update_or_create(
            **lookup,
            defaults={
                "barcode": reference or None,
                "name": name or reference,
                "category": category,
                "unit": unit or "unité",
                "purchase_price": _decimal(row[indexes["purchase_price"]]) if "purchase_price" in indexes else Decimal("0"),
                "wholesale_price": _decimal(row[indexes["wholesale_price"]]) if "wholesale_price" in indexes else Decimal("0"),
                "detail_price": _decimal(row[indexes["detail_price"]]) if "detail_price" in indexes else Decimal("0"),
                "counter_price": _decimal(row[indexes["counter_price"]]) if "counter_price" in indexes else Decimal("0"),
                "default_stock_alert": _decimal(row[indexes["default_stock_alert"]]) if "default_stock_alert" in indexes else Decimal("0"),
                "expiration_date": _date_or_none(row[indexes["expiration_date"]]) if "expiration_date" in indexes else None,
                "shelf_life_days": _int_or_none(row[indexes["shelf_life_days"]]) if "shelf_life_days" in indexes else None,
                "compliance_required": _bool(row[indexes["compliance_required"]]) if "compliance_required" in indexes else False,
            },
        )
        balance, created = StockBalance.objects.get_or_create(
            store=store,
            product=product,
            defaults={
                "quantity": opening_stock,
                "min_stock": product.default_stock_alert,
                "average_cost": product.purchase_price,
            },
        )
        if not created and balance.min_stock is None:
            balance.min_stock = product.default_stock_alert
            balance.save(update_fields=["min_stock", "date_updated"])
        imported += 1

    batch.imported_count = imported
    batch.skipped_count = skipped
    batch.save(update_fields=["imported_count", "skipped_count"])
    return batch
